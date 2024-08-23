import bs4
from langchain_chroma import Chroma
from langchain_community.document_loaders import WebBaseLoader, PlaywrightURLLoader, SeleniumURLLoader

from langchain_core.output_parsers import JsonOutputParser, StrOutputParser
from langchain_core.runnables import RunnablePassthrough

# from langchain_openai import OpenAIEmbeddings
# from langchain_huggingface import HuggingFaceEmbeddings
from langchain_cohere import CohereEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.prompts import PromptTemplate
# import torch
from langchain_community.document_transformers import EmbeddingsRedundantFilter
from langchain.retrievers.document_compressors import DocumentCompressorPipeline
from langchain.retrievers import ContextualCompressionRetriever
from langchain_groq import ChatGroq
from langchain_core.pydantic_v1 import BaseModel, Field
from typing import Any
import re

from googlesearch import search
import pandas as pd
import numpy as np
from dotenv import load_dotenv
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import requests
import asyncio 
import os
import time

# lower down security level
# requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += ":HIGH:!DH:!aNULL"
        
def web_search(keywords, num):
    # keywords = "startup competition 2024"
    num_results = num  # Number of search results to fetch
    df = pd.DataFrame(columns=["Title", "Description", "Url"])
    url_list = []
    i = 0
    for result in search(keywords, num_results=num_results, advanced=True):
        # Process each search result URL
        url = result.url
        url_list.append(result.url)
        title = result.title
        description = result.description
        df.loc[i] = [title, description, url]
        i += 1
    st.write(url_list)
    return url_list
    # Add code to scrape competition information from this website


def web_scrape(url_list):

    loader = WebBaseLoader(
        web_paths=(url_list),
        bs_kwargs=dict(parse_only=bs4.SoupStrainer(name=["p", "div", "h[1-6]"])),
    )

    multiple_docs = loader.load()
    # for i, docs in enumerate(multiple_docs):
    #     if docs.page_content == '': #dynamic web content
    #         st.write("dynamic web content detected")
    #         # loader = PlaywrightURLLoader(urls=url_list[i], remove_selectors=["header", "footer"])
    #         loader =  SeleniumURLLoader(urls=url_list[i])
    #         if loader.load() == []:
    #             pass
    #         else:
    #             multiple_docs[i] == loader.load()
    remove_list = []
    for i, docs in enumerate(multiple_docs):
        if docs.page_content == '':
            remove_list.append(i)
    for i in sorted(remove_list, reverse=True):
        st.write(f"The content from {url_list[i]} cannot be scraped, it will be skipped for information extraction")
        del multiple_docs[i]
        del url_list[i]

    return multiple_docs


def get_chunks_retriever(docs):
    # device = "cuda" if torch.cuda.is_available() else "cpu"
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1500, chunk_overlap=500)
    splits = text_splitter.split_documents([docs])

#     embeddings_model = HuggingFaceEmbeddings(
#         model_name="sentence-transformers/all-mpnet-base-v2",
#         model_kwargs={"device": device},
#     )
    embeddings_model = CohereEmbeddings(
                model="embed-multilingual-v3.0",
                )
    
    vectorstore = Chroma.from_documents(documents=splits, embedding=embeddings_model)
    # Retrieve and generate using the relevant snippets of the blog.
    retriever = vectorstore.as_retriever(search_kwargs={"k": 5})
    # Drop duplicated text chunks
    redundant_filter = EmbeddingsRedundantFilter(embeddings=embeddings_model)

    pipeline_compressor = DocumentCompressorPipeline(
        transformers=[text_splitter, redundant_filter]
    )
    compression_retriever = ContextualCompressionRetriever(
        base_compressor=pipeline_compressor, base_retriever=retriever
    )
    return compression_retriever, vectorstore


def get_comp_chain(retriever, vectorstore):
    llm = ChatGroq(model="llama-3.1-70b-versatile", temperature=0.0)

    template = """Use the following pieces of context to summarize the competition information stated in the question at the end.
    If you don't know the answer, just say that you don't know, don't try to make up an answer.
    Use three sentences maximum and keep the answer as concise as possible for each information.
    {format_instructions}
    
    {context}

    Question: {question}

    Helpful Answer:"""

    parser = JsonOutputParser(pydantic_object=Comp_format)
    custom_rag_prompt = PromptTemplate(
        template=template,
        input_variables=["question"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    rag_chain = (
        {
            "context": retriever | format_docs,
            "question": RunnablePassthrough(),
        }
        | custom_rag_prompt
        | llm
        | parser
    )
    response = rag_chain.invoke(
        "Respond in JSON with keys 'Competition_name', 'Application_deadline', \
        'Region', 'Decription', 'Topics', 'Prize', and 'Eligibility'"
    )
    # Clean up
    vectorstore.delete_collection()

    return response


def get_fund_chain(retriever, vectorstore):
    llm = ChatGroq(model="llama-3.1-70b-versatile", temperature=0.0)

    template = """Use the following pieces of context to summarize the funding opportunity information stated in the question at the end.
    If you don't know the answer, just say that you don't know, don't try to make up an answer.
    Use three sentences maximum and keep the answer as concise as possible for each information.
    {format_instructions}
    
    {context}

    Question: {question}

    Helpful Answer:"""

    parser = JsonOutputParser(pydantic_object=Fund_format)
    custom_rag_prompt = PromptTemplate(
        template=template,
        input_variables=["question"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    rag_chain = (
        {
            "context": retriever | format_docs,
            "question": RunnablePassthrough(),
        }
        | custom_rag_prompt
        | llm
        | parser
    )
    response = rag_chain.invoke(
        "Respond in JSON with keys 'Funding_name', 'Organization', \
        'Application_start', 'Application_end', 'Description', 'Requirements', 'Type' (as investment/project/tender or other types), \
        'Amount', 'Eligibility'"
    )
    # Clean up
    vectorstore.delete_collection()

    return response


class Comp_format(BaseModel):
    Competition_name: Any = Field(description="The name")
    Application_deadline: Any = Field(description="The deadline")
    Region: Any = Field(description="The region")
    Description: Any = Field(description="The description")
    Topics: Any = Field(description="The topics")
    Prize: Any = Field(description="The prize")
    Eligibility: Any = Field(description="The eligibility")


class Fund_format(BaseModel):
    Funding_name: Any = Field(description="The name")
    Organization: Any = Field(description="The organization")
    Application_start: Any = Field(description="The start date")
    Application_end: Any = Field(description="The end date")
    Description: Any = Field(description="The description")
    Requirements: Any = Field(description="The requirements")
    Type: Any = Field(description="The types")
    Amount: Any = Field(description="The amount and prize")
    Eligibility: Any = Field(description="The eligibility")


def format_docs(docs):
    for doc in docs:
        text = re.sub(r"\n+", "\n", doc.page_content)
        text = re.sub(r"\t+", "[t]", text)
        doc.page_content = re.sub(r"\s+", " ", text)

    return "\n\n".join(doc.page_content for doc in docs)


def rag_llm(keywords, num, type):
    results = []
    st.write("Performing web search")
    url_list = web_search(keywords, num)
    st.write("Scraping content")
    num_call = 0
    multiple_docs = web_scrape(url_list) 
    start = time.time()    
    for docs in multiple_docs:
        st.write("Retrieving relevant text chunks")
        if num_call ==6:
            num_call=1 #reset counting
            start = time.time() 
        else:
            num_call+=num_call
        # st.write(docs)
        duration = time.time()-start
        if num_call== 6 and duration < 60: #reached cohere embedding rate limit (5 calls/minute)
            time.sleep(70-duration)
        compression_retriever, vectorstore = get_chunks_retriever(docs)
        if type == "competition":
            results.append(get_comp_chain(compression_retriever, vectorstore))
        else:  # == "funding opportunity"
            results.append(get_fund_chain(compression_retriever, vectorstore))
        st.write("Results generated for this record")

    return results, url_list


def sequential_calls(keywords, num, type):
    results, url_list = rag_llm(keywords, num, type)
    # for result in results:
    #     print(result)
    #     print("\n")
    return results, url_list


def results_processing(results, url_list):
    processed_results = []

    # Convert json to list format where result with multiple records are splited into single record
    for i, result in enumerate(results):
        if not isinstance(result, list):  # only one competiiton recorded in this data
            result["Link"] = url_list[i]
            processed_results.append(result)
        else:  # more than 1 competition recorded in this data
            for record in result:
                record["Link"] = url_list[i]
                processed_results.append(record)

    return processed_results


# Convert JSON data to a DataFrame
def results_to_excel(processed_results, type):
    df = pd.DataFrame(processed_results)
    time = datetime.now()
    time = time.strftime("%Y%m%d_%H%M%S")

    # Specify the Excel file path
    if type == "competition":
        excel_file_path = f"competitions_{time}.xlsx"
    else:
        excel_file_path = f"funding_{time}.xlsx"

    # Write the DataFrame to an Excel file
    df.to_excel(excel_file_path, index=False)

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file_path)
    worksheet = workbook.active

    # Adjust column width before wrapping text
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            # Calculate max_length for autofit
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        # Adjust column width based on the max length
        adjusted_width = max_length + 2  # Add some extra space
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Apply wrap text to all cells and set borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in worksheet.iter_rows(
        min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column
    ):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border

    # Define table range and create a table
    table_range = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"
    table = Table(displayName="DataTable", ref=table_range)

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    worksheet.add_table(table)

    # Save the adjusted workbook
    workbook.save(excel_file_path)

    print(
        f"DataFrame successfully written to {excel_file_path} with borders, wrapped text, and as a table."
    )

    return excel_file_path
